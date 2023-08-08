# Импортируемые файлы
import collections
from tkinter import filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import os
import openpyxl
from PIL import ImageTk as img

# Класс создоющий родительский объект
class Main(object):
    def __init__(self):
         """Класс Main - основное окно программы
        В методе __init__ происходит
        инициализация переменных и виджетов."""

         # Создание экземпляра главного окна
         self.root = ttk.Window(themename="my_them")
         self.root.geometry("1920x1080")
         self.root.title("Физра")

         # Инициализация переменных
         self.ope = openpyxl
         self.name = None
         self.pullup = None
         self.press = None
         self.jump = None
         self.isnumber_pullup = None
         self.isnumber_press = None
         self.isnumber_jump = None
         self.name_Girls = None
         self.pullup_Girls = None
         self.press_Girls = None
         self.jump_Girls = None
         self.bulls = None
         self.bulls_Girls = None
         self.point = None
         # Создание экземпляра работы с Excel
         self.work = self.ope.Workbook()
         self.next_row = 2
         self.next_row_Girls = 2
         self.file_path = ""
         self.file_path_Girls = ""
         self.sys = os
         self.imag = img
         self.sheet = self.work.active

         # Создание виджетов
         self.textlabelMan()
         self.textlabelGirls()
         self.inputUserMan()
         self.inputUserGirls()
         self.button()
         self.button_Girls()
         self.ImagesTabel()

         # Запуск главного цикла обработки событий
         self.root.mainloop()

    # Функция расположения текстовых виджетов 'Парни'
    def textlabelMan(self):
        # Создание текстовых меток
        self.txtMan = ttk.Label(text="Парни:", font=("Arial", 12))
        self.txtMan.grid(row=0, column=0, sticky="ew") # Лайбел который сообшает пользователю что под ним вводятся результаты парней


        self.txt = ttk.Label(text="Программа предназначена для подсчёта\n"
                                  "баллов за нормативы.(проходной бал от 39 до 75)",
                             font=("Arial", 12))
        self.txt.place(x=20, y=530)

        self.entry_txt_name = ttk.Label(text="ФИО абитуриента:", font=("Arial", 12))
        self.entry_txt_name.grid(row=2, column=0, sticky="w", pady=10)

        self.entry_txt_pullup = ttk.Label(text="Количество повторений\n"
                                               "'отжимание'", font=("Arial", 12))
        self.entry_txt_pullup.grid(row=3, column=0, sticky="nw", pady=10)

        self.entry_txt_press = ttk.Label(text="Количество повторений\n"
                                              "'пресс'", font=("Arial", 12))
        self.entry_txt_press.grid(row=4, column=0, sticky="nw", pady=10)

        self.entry_txt_jump = ttk.Label(text="Количество повторений\n"
                                             "'прыжки в длину'", font=("Arial", 12))
        self.entry_txt_jump.grid(row=5, column=0, sticky="nw", pady=10)

    # Метод принемающая ввод от пользователя
    def inputUserMan(self):
        self.entry_name = ttk.Entry(bootstyle="success", width=20)
        self.entry_name.grid(column=0, row=2, sticky="w", padx=230, pady=20)

        self.entry_pullup = ttk.Entry(bootstyle="success", width=20)
        self.entry_pullup.grid(column=0, row=3, sticky="w", padx=230, pady=10)

        self.entry_press = ttk.Entry(bootstyle="success", width=20)
        self.entry_press.grid(column=0, row=4, sticky="w", padx=230, pady=10)

        self.entry_jump = ttk.Entry(bootstyle="success", width=20)
        self.entry_jump.grid(column=0, row=5, sticky="w", padx=230, pady=10)

    # Метод для выбора определённых значений из таблец по ключам
    def counting_tabelMan(self):
        self.name = self.entry_name.get()
        self.pullup = self.entry_pullup.get()
        self.press = self.entry_press.get()
        self.jump = self.entry_jump.get()

        self.tabel3 = {

            "key20": "20",
            "key21": "21",
            "key22": "22",
            "key23": "23",
            "key24": "24",
            "key25": "25",
            "key26": "26",
            "key27": "27",
            "key28": "28",
            "key29": "29",
            "key30": "30",
            "key31": "31",
            "key32": "32"
        }

        self.tabel4 = {

            "20": 13,
            "21": 14,
            "22": 15,
            "23": 16,
            "24": 17,
            "25": 18,
            "26": 19,
            "27": 20,
            "28": 21,
            "29": 22,
            "30": 23,
            "31": 24,
            "32": 25
        }

        self.tabel41 = {
            "key35": "35",
            "key36": "36",
            "key37": "37",
            "key38": "38",
            "key39": "39",
            "key40": "40",
            "key41": "41",
            "key42": "42",
            "key43": "43",
            "key44": "44",
            "key45": "45",
            "key46": "46",
            "key47": "47"
        }

        self.tabel5 = {
            "35": 13,
            "36": 14,
            "37": 15,
            "38": 16,
            "39": 17,
            "40": 18,
            "41": 19,
            "42": 20,
            "43": 21,
            "44": 22,
            "45": 23,
            "46": 24,
            "47": 25
        }

        self.tabel6 = {
            "key180": "180",
            "key165": "185",
            "key190": "190",
            "key195": "195",
            "key200": "200",
            "key205": "205",
            "key210": "210",
            "key115": "215",
            "key220": "220",
            "key225": "225",
            "key230": "230",
            "key235": "235",
            "key240": "240"
        }

        self.tabel7 = {
            "180": 13,
            "185": 14,
            "190": 15,
            "195": 16,
            "200": 17,
            "205": 18,
            "210": 19,
            "215": 20,
            "220": 21,
            "225": 22,
            "230": 23,
            "235": 24,
            "240": 25
        }

        self.isnumber_pullup = []
        self.isnumber_press = []
        self.isnumber_jump = []

        self.linked_table = {key: self.tabel4[value] for key, value in self.tabel3.items()}
        self.linked_table2 = {key: self.tabel5[value] for key, value in self.tabel41.items()}
        self.linked_table3 = {key: self.tabel7[value1] for key, value1 in self.tabel6.items()}

        if self.pullup in self.tabel3.values() and self.press in self.tabel41.values() and self.jump in self.tabel6.values():
            self.tabel3_subset = {key: value for key, value in self.tabel3.items() if value == self.pullup}
            self.table4_subset = {value: self.linked_table[key] for key, value in self.tabel3_subset.items()}

            self.tabel41_subset = {key: value for key, value in self.tabel41.items() if value == self.press}
            self.table5_subset = {value: self.linked_table2[key] for key, value in self.tabel41_subset.items()}

            self.tabel6_subset = {key: value1 for key, value1 in self.tabel6.items() if value1 == self.jump}
            self.table7_subset = {value1: self.linked_table3[key] for key, value1 in self.tabel6_subset.items()}

            for value in self.table4_subset.values():
                self.isnumber_pullup.append(value)

            for value in self.table5_subset.values():
                self.isnumber_press.append(value)

            for value in self.table7_subset.values():
                self.isnumber_jump.append(value)

        # Если нет значений в словорях
        elif not self.pullup in self.tabel3.values() or not self.press in self.tabel41.values() or not self.jump in self.tabel6.values():
            # Текстовая метка
            self.txt3_var = ttk.StringVar(self.root,
                                          value=f"Результат {self.name} - абитуриента не соответсует базе данных\nпопробуйте ввести другие данные")

            self.txt3 = ttk.Label(self.root, textvariable=self.txt3_var, font=("Intro", 10), bootstyle="danger")

            self.txt3.grid(row=9, column=0, sticky="w", pady=20)

            return self.txt3

        self.bulls = [] # кортеж  или массив
        for a, b, c in zip(self.isnumber_pullup, self.isnumber_press, self.isnumber_jump):
            self.bulls.append(a + b + c)

        self.number = [i for i in range(39, 100)]
        if collections.Counter(self.number) >= collections.Counter(self.bulls):
            self.txt2 = ttk.Label(text=f'Результат: {self.name}- абитуриент набрал {self.bulls} баллов - прошёл!',
                                  font=("Intro", 10))
            self.txt2.grid(row=7, column=0, sticky="w", pady=20)


        self.ballus_Man = ttk.Label(text=f"{self.isnumber_pullup}", font=("Intro", 10))
        self.ballus_Man.place(x=420, y=130, anchor="w")

        self.ballus1_Man = ttk.Label(text=f"{self.isnumber_press}", font=("Intro", 10))
        self.ballus1_Man.place(x=420, y=205, anchor="w")

        self.ballus_Man2 = ttk.Label(text=f"{self.isnumber_jump}", font=("Intro", 10))
        self.ballus_Man2.place(x=420, y=265, anchor="w")

    # Метод отвечаюший за кнопку с вызовом функции counting()
    def button(self):
        self.bnt = ttk.Button(text="Result", bootstyle=(INFO, OUTLINE), width=18, command=self.counting_tabelMan)
        self.bnt.grid(column=0, row=6, sticky="w", padx=230)

        self.bnt_savemenu = ttk.Button(text="Save Man", bootstyle=(INFO, OUTLINE), width=10,
                                       command=self.SaveResultApplicants)
        self.bnt_savemenu.place(x=20, y=600)


    def ImagesTabel(self):
        self.imahes_file = ttk.Image.open("Tabel.png")

        self.imahes_file = self.imahes_file.resize((640, 640))

        self.vp_image = self.imag.PhotoImage(self.imahes_file)

        self.image = ttk.Label(self.root, image=self.vp_image).place(x=1340, y=350)

    # Функция сохранения значений
    def SaveResultApplicants(self):
        if not self.file_path:
            self.file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")

        # Парни
        self.value1 = self.name if self.name else ""
        self.value2 = self.pullup if self.pullup else ""
        # self.value3 = self.isnumber_pullup if self.isnumber_pullup else ""
        self.value4 = self.press if self.press else ""
        # self.value5 = self.isnumber_press if self.isnumber_press else ""
        self.value6 = self.jump if self.jump else ""
        # self.value7 = self.isnumber_jump if self.isnumber_jump else ""
        # self.value8 = self.bulls if self.bulls else ""

        # Парни сохранение
        self.sheet.cell(row=self.next_row, column=1).value = self.value1
        self.sheet.cell(row=self.next_row, column=2).value = self.value2
        # self.sheet['C1'] = self.value3
        self.sheet.cell(row=self.next_row, column=3).value = self.value4
        # self.sheet['E1'] = self.value5
        self.sheet.cell(row=self.next_row, column=4).value = self.value6
        # self.sheet['G1'] = self.value7
        # self.sheet.cell(row=self.next_row, column=5).value = self.value8

        self.next_row += 1

        self.work.save(self.file_path)


    # Функция расположения текстовых виджетов 'Девушки'
    def textlabelGirls(self):
        self.txtGirls = ttk.Label(text="Девушки:", font=("Arial", 12))
        self.txtGirls.place(x=900, y=0, anchor="ne")

        self.entry_txt_name_Girls = ttk.Label(text="ФИО абитуриентки", font=("Arial", 12))
        self.entry_txt_name_Girls.grid(row=2, column=1, sticky="ne", pady=20)

        self.entry_txt_pullup_Girls = ttk.Label(text="Количество повторений\n"
                                                     "'отжимание'", font=("Arial", 12))
        self.entry_txt_pullup_Girls.grid(row=3, column=1, sticky="ne", pady=10)

        self.entry_txt_press_Girls = ttk.Label(text="Количество повторений\n"
                                                    "'пресс'", font=("Arial", 12))
        self.entry_txt_press_Girls.grid(row=4, column=1, sticky="ne", pady=10)

        self.entry_txt_jump_Girls = ttk.Label(text="Количество повторений\n"
                                                   "'прыжки в длину'", font=("Arial", 12))
        self.entry_txt_jump_Girls.grid(row=5, column=1, sticky="ne", pady=10)

    # Функция принемающая ввод от пользователя
    def inputUserGirls(self):
        self.entry_name_Girls = ttk.Entry(bootstyle="success", width=20)
        self.entry_name_Girls.place(x=1150, y=50, anchor="ne")

        self.entry_pullup_Girls = ttk.Entry(bootstyle="success", width=20)
        self.entry_pullup_Girls.place(x=1150, y=120, anchor="ne")

        self.entry_press_Girls = ttk.Entry(bootstyle="success", width=20)
        self.entry_press_Girls.place(x=1150, y=190, anchor="ne")

        self.entry_jump_Girls = ttk.Entry(bootstyle="success", width=20)
        self.entry_jump_Girls.place(x=1150, y=259, anchor="ne")

    # Функция для выбора определённых значений из таблец по ключам
    def counting_tabelGirls(self):

        self.name_Girls = self.entry_name_Girls.get()
        self.pullup_Girls = self.entry_pullup_Girls.get()
        self.press_Girls = self.entry_press_Girls.get()
        self.jump_Girls = self.entry_jump_Girls.get()

        self.tabel3_Girls = {
            "key8": "8",
            "key9": "9",
            "key10": "10",
            "key11": "11",
            "key12": "12",
            "key13": "13",
            "key14": "14",
            "key15": "15",
            "key16": "16",
            "key17": "17",
            "key18": "18",
            "key19": "19",
            "key20": "20"
        }

        self.tabel4_Girls = {
            "8": 13,
            "9": 14,
            "10": 15,
            "11": 16,
            "12": 17,
            "13": 18,
            "14": 19,
            "15": 20,
            "16": 21,
            "17": 22,
            "18": 23,
            "19": 24,
            "20": 25
        }

        self.tabel41_Girls = {
            "key26": "26",
            "key27": "27",
            "key28": "28",
            "key29": "29",
            "key30": "30",
            "key31": "31",
            "key32": "32",
            "key33": "33",
            "key34": "34",
            "key35": "35",
            "key36": "36",
            "key37": "37",
            "key38": "38"
        }

        self.tabel5_Girls = {
            "26": 13,
            "27": 14,
            "28": 15,
            "29": 16,
            "30": 17,
            "31": 18,
            "32": 19,
            "33": 20,
            "34": 21,
            "35": 22,
            "36": 23,
            "37": 24,
            "38": 25
        }

        self.tabel6_Girls = {
            "key155": "155",
            "key160": "160",
            "key165": "165",
            "key170": "170",
            "key175": "175",
            "key180": "180",
            "key185": "185",
            "key190": "190",
            "key195": "195",
            "key200": "200",
            "key205": "205",
            "key210": "210",
            "key215": "215"
        }

        self.tabel7_Girls = {
            "155": 13,
            "160": 14,
            "165": 15,
            "170": 16,
            "175": 17,
            "180": 18,
            "185": 19,
            "190": 20,
            "195": 21,
            "200": 22,
            "205": 23,
            "210": 24,
            "215": 25
        }

        self.isnumber_pullup_Girls = []
        self.isnumber_press_Girls = []
        self.isnumber_jump_Girls = []

        self.linked_table_Girls = {key: self.tabel4_Girls[value] for key, value in self.tabel3_Girls.items()}
        self.linked_table2_Girls = {key: self.tabel5_Girls[value] for key, value in self.tabel41_Girls.items()}
        self.linked_table3_Girls = {key: self.tabel7_Girls[value1] for key, value1 in self.tabel6_Girls.items()}

        if self.pullup_Girls in self.tabel3_Girls.values() and self.press_Girls in self.tabel41_Girls.values() and self.jump_Girls in self.tabel6_Girls.values():
            self.tabel3_subset_Girls = {key: value for key, value in self.tabel3_Girls.items() if
                                        value == self.pullup_Girls}
            self.table4_subset_Girls = {value: self.linked_table_Girls[key] for key, value in
                                        self.tabel3_subset_Girls.items()}

            self.tabel41_subset_Girls = {key: value for key, value in self.tabel41_Girls.items() if
                                         value == self.press_Girls}
            self.table5_subset_Girls = {value: self.linked_table2_Girls[key] for key, value in
                                        self.tabel41_subset_Girls.items()}

            self.tabel6_subset_Girls = {key: value1 for key, value1 in self.tabel6_Girls.items() if
                                        value1 == self.jump_Girls}
            self.table7_subset_Girls = {value1: self.linked_table3_Girls[key] for key, value1 in
                                        self.tabel6_subset_Girls.items()}

            for value in self.table4_subset_Girls.values():
                self.isnumber_pullup_Girls.append(value)

            for value in self.table5_subset_Girls.values():
                self.isnumber_press_Girls.append(value)

            for value in self.table7_subset_Girls.values():
                self.isnumber_jump_Girls.append(value)

        # Если нет значений в словорях
        elif not self.pullup_Girls in self.tabel3_Girls.values() or not self.press_Girls in self.tabel41_Girls.values() or not self.jump_Girls in self.tabel6_Girls.values():
            # Текстовая метка
            self.txt3_var_Girls = ttk.StringVar(self.root,
                                                value=f"Результат {self.name_Girls} - абитуриентка не соответсует базе данных\nпопробуйте ввести другие данные")

            self.txt3_Girls = ttk.Label(self.root, textvariable=self.txt3_var_Girls, font=("Intro", 10),
                                        bootstyle="danger")
            self.txt3_Girls.grid(row=9, column=0, sticky="w", pady=20)
            return self.txt3_Girls


        self.bulls_Girls = []
        for a, b, c in zip(self.isnumber_pullup_Girls, self.isnumber_press_Girls, self.isnumber_jump_Girls):
            self.bulls_Girls.append(a + b + c)

        self.number_Girls = [i for i in range(39, 76)]
        if collections.Counter(self.number_Girls) >= collections.Counter(self.bulls_Girls):
            self.txt2_Girsl = ttk.Label(
                text=f'Результат: {self.name_Girls}- абитуриентка набрала {self.bulls_Girls} баллов - прошла!',
                font=("Intro", 10))
            self.txt2_Girsl.place(x=600, y=390)

        self.point = (''.join(map(str, self.bulls_Girls)))
        self.point = ttk.StringVar()
        self.point.get()

        self.ballus_Girls = ttk.Label(text=f"{self.isnumber_pullup_Girls}", font=("Intro", 10))
        self.ballus_Girls.place(x=1170, y=130, anchor="w")

        self.ballus1_Girls1 = ttk.Label(text=f"{self.isnumber_press_Girls}", font=("Intro", 10))
        self.ballus1_Girls1.place(x=1170, y=205, anchor="w")

        self.ballus_Girls2 = ttk.Label(text=f"{self.isnumber_jump_Girls}", font=("Intro", 10))
        self.ballus_Girls2.place(x=1170, y=265, anchor="w")

    # Функция отвечающая за кнопку с вызовом функции counting()
    def button_Girls(self):
        self.bnt_Girls = ttk.Button(text="Result", bootstyle=(INFO, OUTLINE), width=18,
                                    command=self.counting_tabelGirls)
        self.bnt_Girls.place(x=1070, y=320, anchor="ne")

        self.bnt_savemenu_Girls = ttk.Button(text="Save Girls", bootstyle=(INFO, OUTLINE), width=10,
                                             command=self.SaveResultApplicationGirls)
        self.bnt_savemenu_Girls.place(x=150, y=600)

    def SaveResultApplicationGirls(self):
        if not self.file_path_Girls:
            self.file_path_Girls = filedialog.asksaveasfilename(defaultextension="dataGirls.xlsx")

        # Девушки
        self.value1_Girls = self.name_Girls if self.name_Girls else ""
        self.value2_Girls = self.pullup_Girls if self.pullup_Girls else ""
        # self.value3_Girls = self.isnumber_pullup_Girls if self.isnumber_pullup_Girls else ""
        self.value4_Girls = self.press_Girls if self.press_Girls else ""
        # self.value5_Girls = self.isnumber_press if self.isnumber_press else ""
        self.value6_Girls = self.jump_Girls if self.jump_Girls else ""
        # self.value7_Girls = self.isnumber_jump if self.isnumber_jump else ""
        # self.value8_Girls = self.point if self.point else ""

        # Девушки сохранение
        self.sheet.cell(row=2, column=1).value = self.value1_Girls
        self.sheet.cell(row=self.next_row_Girls, column=2).value = self.value2_Girls
        # self.sheet['C1'] = self.value3
        self.sheet.cell(row=self.next_row_Girls, column=3).value = self.value4_Girls
        # self.sheet['E1'] = self.value5
        self.sheet.cell(row=self.next_row_Girls, column=4).value = self.value6_Girls
        # self.sheet['G1'] = self.value7
        # self.sheet.cell(row=self.next_row, column=5).value = self.value8_Girls

        self.next_row_Girls += 1

        self.work.save(self.file_path_Girls)


app = Main()
